import os
import json
import smtplib
from email.message import EmailMessage
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv

# Load environment variables securely
load_dotenv()

SENDER_EMAIL = os.getenv("SENDER_EMAIL")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
RECIPIENT_EMAIL = os.getenv("RECIPIENT_EMAIL")

HISTORY_FILE = "history.json"
REPORT_NAME = f"Monthly_Price_Report_{datetime.now().strftime('%Y_%B')}.xlsx"

def process_history_data():
    if not os.path.exists(HISTORY_FILE):
        return None
        
    with open(HISTORY_FILE, "r") as f:
        try:
            history = json.load(f)
        except:
            return None
            
    if not history:
        return None
        
    # Group data by property
    stats = {}
    
    # Sort history by date chronologically
    history.sort(key=lambda x: x["date"])
    
    for daily_entry in history:
        date_str = daily_entry["date"]
        for prop in daily_entry["data"]:
            name = prop["name"]
            price = prop["price"]
            prop_type = prop.get("type", "hotel").upper()
            
            if name not in stats:
                stats[name] = {
                    "type": prop_type,
                    "prices": [],
                    "start_price": price,
                    "end_price": price,
                }
            
            stats[name]["prices"].append(price)
            stats[name]["end_price"] = price # Overwritten sequentially, ends at the latest
            
    # Calculate final aggregates
    for name, data in stats.items():
        prices = data["prices"]
        data["min_price"] = min(prices)
        data["max_price"] = max(prices)
        data["avg_price"] = sum(prices) / len(prices)
        data["monthly_variation"] = data["end_price"] - data["start_price"]
        
    return stats

def create_excel_report(stats):
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Statistics"
    
    # Headers
    headers = [
        "Property Name", "Type", "Start Price (USD)", "End Price (USD)", 
        "Monthly Variation (USD)", "Min Price (USD)", "Max Price (USD)", "Avg Price (USD)"
    ]
    
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20
        
    # Data Rows
    row_num = 2
    for name, data in stats.items():
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=data["type"])
        ws.cell(row=row_num, column=3, value=data["start_price"])
        ws.cell(row=row_num, column=4, value=data["end_price"])
        
        # Color variation
        var_cell = ws.cell(row=row_num, column=5, value=data["monthly_variation"])
        if data["monthly_variation"] > 0:
            var_cell.font = Font(color="FF0000", bold=True) # Red if price went up
        elif data["monthly_variation"] < 0:
            var_cell.font = Font(color="00B050", bold=True) # Green if price went down
            
        ws.cell(row=row_num, column=6, value=data["min_price"])
        ws.cell(row=row_num, column=7, value=data["max_price"])
        ws.cell(row=row_num, column=8, value=round(data["avg_price"], 2))
        
        row_num += 1
        
    wb.save(REPORT_NAME)
    return REPORT_NAME

def send_email(attachment_path):
    if not SENDER_EMAIL or not SENDER_PASSWORD or not RECIPIENT_EMAIL:
        print("Missing email credentials in .env file. Skipping email dispatch.")
        return False
        
    msg = EmailMessage()
    msg['Subject'] = f"Monthly Hotel & Apartment Price Statistics - {datetime.now().strftime('%B %Y')}"
    msg['From'] = SENDER_EMAIL
    msg['To'] = RECIPIENT_EMAIL
    
    content = f"""
    Hello,
    
    Please find attached the automated monthly statistics report for the HCMC properties.
    This report contains the price variations (increases and decreases) tracked over the course of the month.
    
    Generated automatically by your system.
    """
    msg.set_content(content)
    
    with open(attachment_path, 'rb') as f:
        file_data = f.read()
        
    msg.add_attachment(file_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=attachment_path)
    
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
            smtp.send_message(msg)
        print(f"Report successfully sent to {RECIPIENT_EMAIL}")
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

def generate_and_send_report():
    print("Gathering statistics from history...")
    stats = process_history_data()
    if not stats:
        print("No historical data found to generate report.")
        return
        
    print("Generating Excel report...")
    report_path = create_excel_report(stats)
    print(f"Excel report saved as {report_path}")
    
    print("Sending email...")
    send_email(report_path)

if __name__ == "__main__":
    generate_and_send_report()
